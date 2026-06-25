"""
invoice_parser.py
Phase 8.1 — Invoice Upload Channel
อ่านไฟล์ PDF / Image / Excel / CSV → structured JSON

รองรับ:
- PDF (text)   → pdfplumber
- PDF (scanned)→ ส่งให้ Claude Vision
- Image        → ส่งให้ Claude Vision
- Excel/XLSX   → openpyxl / pandas
- CSV          → pandas
"""

import io
import os
import json
import base64
from typing import Optional

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")


# ─── Detect file type ────────────────────────────────────────────────────────

def detect_file_type(filename: str, content: bytes) -> str:
    """คืน PDF_TEXT / PDF_SCAN / IMAGE / EXCEL / CSV"""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("jpg", "jpeg", "png", "webp", "gif", "bmp"):
        return "IMAGE"
    if ext in ("xlsx", "xls"):
        return "EXCEL"
    if ext == "csv":
        return "CSV"
    if ext == "pdf" or content[:4] == b"%PDF":
        # ลอง extract text ดูก่อน ถ้ามีข้อความพอ = PDF_TEXT ไม่งั้น = PDF_SCAN
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text = " ".join(
                    (p.extract_text() or "") for p in pdf.pages[:3]
                )
            return "PDF_TEXT" if len(text.strip()) > 100 else "PDF_SCAN"
        except Exception:
            return "PDF_SCAN"
    return "UNKNOWN"


# ─── Extract raw text ─────────────────────────────────────────────────────────

def extract_text_from_pdf(content: bytes) -> str:
    """pdfplumber → raw text ทุกหน้า"""
    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                # ลอง extract table ด้วย
                tables = page.extract_tables() or []
                for table in tables:
                    for row in table:
                        if row:
                            parts.append(" | ".join(str(c or "") for c in row))
                if text:
                    parts.append(text)
        return "\n".join(parts)
    except ImportError:
        return ""
    except Exception as e:
        return f"[PDF parse error: {e}]"


def extract_text_from_excel(content: bytes, filename: str) -> str:
    """openpyxl (raw cell values) + pandas fallback — คืน text ที่ Claude อ่านได้ดี"""
    if filename.lower().endswith(".csv"):
        try:
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content), encoding="utf-8", errors="replace")
            return df.to_string(index=False)
        except Exception as e:
            return f"[CSV parse error: {e}]"

    # Excel — อ่านด้วย openpyxl แบบ raw (ได้ทุก cell รวม merged)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(v).strip().replace("\n", " ") if v is not None else "" for v in row]
                # ข้ามแถวว่างทั้งหมด
                if any(v for v in row_vals):
                    lines.append("\t".join(row_vals))
        return "\n".join(lines)
    except ImportError:
        pass
    except Exception as e:
        pass

    # Fallback pandas
    try:
        import pandas as pd
        xl = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None)
        parts = []
        for name, df in xl.items():
            parts.append(f"=== Sheet: {name} ===")
            parts.append(df.fillna("").to_string(index=False, header=False))
        return "\n".join(parts)
    except Exception as e:
        return f"[Excel parse error: {e}]"


def image_to_base64(content: bytes) -> str:
    return base64.b64encode(content).decode()


# ─── Claude Vision / Text extraction ─────────────────────────────────────────

EXTRACTION_PROMPT = """You are an expert customs document reader. Extract ALL information from this commercial invoice.

Return a JSON object with this exact structure:
{
  "invoice_no": "string or null",
  "invoice_date": "YYYY-MM-DD or null",
  "seller_name": "string or null",
  "seller_country": "ISO 2-letter code or null",
  "buyer_name": "string or null",
  "buyer_country": "ISO 2-letter code or null",
  "incoterms": "FOB/CIF/EXW/DDP/etc or null",
  "currency": "USD/THB/CNY/EUR/etc or null",
  "items": [
    {
      "line_no": 1,
      "description": "exact product description from invoice",
      "hs_code_declared": "HS code if shown, else null",
      "qty": number or null,
      "unit": "KG/PCS/CARTON/CBM/L/etc or null",
      "unit_price": number or null,
      "line_value": number or null,
      "country_origin": "ISO 2-letter or null",
      "marks_numbers": "string or null"
    }
  ]
}

Rules:
- Extract ALL line items, do not skip any
- Use null for missing fields, never guess
- description must be verbatim from the document
- Return ONLY the JSON, no explanation"""


async def extract_with_claude_vision(content: bytes, file_type: str) -> dict:
    """ส่งรูปภาพ/PDF scan ให้ Claude Vision อ่าน"""
    if not ANTHROPIC_API_KEY:
        return {"error": "ANTHROPIC_API_KEY not set"}
    try:
        import httpx

        media_type = "image/jpeg"
        if file_type == "PDF_SCAN":
            try:
                from pdf2image import convert_from_bytes
                images = convert_from_bytes(content, first_page=1, last_page=3)
                img_io = io.BytesIO()
                images[0].save(img_io, format="JPEG")
                content = img_io.getvalue()
                media_type = "image/jpeg"
            except ImportError:
                media_type = "application/pdf"

        b64 = image_to_base64(content)
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 4096,
                    "messages": [{
                        "role": "user",
                        "content": [
                            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                            {"type": "text", "text": EXTRACTION_PROMPT}
                        ]
                    }]
                }
            )
            resp.raise_for_status()
            data = resp.json()
        raw = data["content"][0]["text"].strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except Exception as e:
        return {"error": str(e)}


async def extract_with_claude_text(raw_text: str) -> dict:
    """ส่ง raw text ให้ Claude สกัด structured JSON"""
    if not ANTHROPIC_API_KEY:
        return {"error": "ANTHROPIC_API_KEY not set"}
    try:
        import httpx
        prompt = EXTRACTION_PROMPT + f"\n\nDocument text:\n{raw_text[:8000]}"
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 4096,
                    "messages": [{"role": "user", "content": prompt}]
                }
            )
            resp.raise_for_status()
            data = resp.json()
        raw = data["content"][0]["text"].strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except Exception as e:
        return {"error": str(e)}


# ─── Excel Compiler (อ่าน cell → clean text → Claude) ───────────────────────

def _compile_excel_text(content: bytes, filename: str) -> str:
    """อ่านทุก cell → format เป็น pipe-delimited text"""
    if filename.lower().endswith(".csv"):
        try:
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content), encoding="utf-8", errors="replace")
            return df.to_string(index=False)
        except Exception as e:
            return f"[CSV error: {e}]"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(v).strip().replace("\n", " ") if v is not None else "" for v in row]
            if any(c for c in cells):
                lines.append(" | ".join(cells))
        return "\n".join(lines)
    except Exception as e:
        return f"[Excel read error: {e}]"


def _parse_compiled_text(text: str) -> dict:
    """
    Parse pipe-delimited text จาก _compile_excel_text
    หา header row ที่มี description/hs code/qty แล้ว extract items โดยตรง
    ไม่ต้องพึ่ง Claude — ทำงานบน text แทน openpyxl cells
    """
    import re

    result = {"items": []}
    lines = text.splitlines()

    # Keywords สำหรับ header detection (case-insensitive)
    DESC_KW = ("description", "goods", "commodity", "item", "product", "สินค้า")
    HS_KW   = ("hs code", "hscode", "hs_code", "tariff")
    QTY_KW  = ("qty", "quantity", "จำนวน")
    PRICE_KW = ("unit price", "price", "ราคา")
    AMOUNT_KW = ("amount", "value", "total", "มูลค่า")
    ORIGIN_KW = ("origin",)

    # สกัด meta จาก text ทั้งหมด
    inv_m = re.search(r'(INV[-\s]?[\w\-]+)', text, re.IGNORECASE)
    result["invoice_no"] = inv_m.group(1) if inv_m else None

    date_m = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2})', text)
    result["invoice_date"] = date_m.group(1) if date_m else None

    inco_m = re.search(r'\b(CIF|FOB|EXW|DDP|CFR|DAP|FCA)\b', text, re.IGNORECASE)
    result["incoterms"] = inco_m.group(1).upper() if inco_m else None

    cur_m = re.search(r'\b(USD|THB|CNY|EUR|GBP|JPY)\b', text, re.IGNORECASE)
    result["currency"] = cur_m.group(1).upper() if cur_m else "USD"

    result["seller_country"] = "CN" if "china" in text.lower() else None
    result["buyer_country"] = "TH" if ("thailand" in text.lower() or "bangkok" in text.lower()) else None

    # หา header line
    header_idx = -1
    col_map = {}
    for i, line in enumerate(lines):
        cols = [c.strip().lower() for c in line.split("|")]
        desc_ci = next((ci for ci, v in enumerate(cols) if any(kw in v for kw in DESC_KW)), None)
        if desc_ci is not None:
            col_map["desc"] = desc_ci
            for ci, v in enumerate(cols):
                if any(kw in v for kw in HS_KW):    col_map["hs"] = ci
                if any(kw in v for kw in QTY_KW):   col_map["qty"] = ci
                if any(kw in v for kw in PRICE_KW): col_map["price"] = ci
                if any(kw in v for kw in AMOUNT_KW): col_map.setdefault("amount", ci)
                if any(kw in v for kw in ORIGIN_KW): col_map["origin"] = ci
            # ถัดจาก qty = unit
            if "qty" in col_map:
                unit_ci = col_map["qty"] + 1
                if unit_ci not in col_map.values():
                    col_map["unit"] = unit_ci
            header_idx = i
            break

    result["_debug"] = f"lines={len(lines)} header_idx={header_idx} col_map={col_map} first_lines={[l[:60] for l in lines[:5]]}"
    if header_idx == -1:
        return result  # ไม่เจอ header → fallback

    # Extract items
    items = []
    SKIP_KW = ("total", "subtotal", "freight", "insurance", "remark", "signature", "authorized")
    for line in lines[header_idx + 1:]:
        cols_raw = [c.strip() for c in line.split("|")]
        if col_map["desc"] >= len(cols_raw):
            continue
        desc = cols_raw[col_map["desc"]]
        if not desc or any(kw in desc.lower() for kw in SKIP_KW):
            continue

        def get(key):
            ci = col_map.get(key)
            return cols_raw[ci].strip() if ci is not None and ci < len(cols_raw) else None

        def to_float(v):
            try:
                if not v:
                    return None
                cleaned = str(v).replace(",", "").replace("$", "").replace("฿", "").replace("€", "").replace("¥", "").strip()
                return float(cleaned) if cleaned else None
            except Exception:
                return None

        qty_val = to_float(get("qty"))
        unit_price_val = to_float(get("price"))
        line_value_val = to_float(get("amount"))
        # fallback: คำนวณ qty × unit_price ถ้า line_value ไม่มี
        if line_value_val is None and qty_val and unit_price_val:
            line_value_val = round(qty_val * unit_price_val, 2)

        items.append({
            "line_no": len(items) + 1,
            "description": desc,
            "hs_code_declared": get("hs"),
            "qty": qty_val,
            "unit": get("unit"),
            "unit_price": unit_price_val,
            "line_value": line_value_val,
            "country_origin": get("origin") or result.get("seller_country"),
            "marks_numbers": None,
        })

    result["items"] = items
    return result


# ─── Smart Excel Parser ──────────────────────────────────────────────────────

# keywords ที่บ่งบอกว่าแถวนี้คือ header ของตาราง items
_ITEM_HEADER_KEYWORDS = {
    "description", "goods", "commodity", "item", "product",
    "สินค้า", "รายการ", "คำบรรยาย"
}
_HS_KEYWORDS = {"hs", "tariff", "hscode", "hs code", "hs_code"}
_QTY_KEYWORDS = {"qty", "quantity", "จำนวน"}
_PRICE_KEYWORDS = {"price", "unit price", "ราคา"}
_AMOUNT_KEYWORDS = {"amount", "total", "value", "มูลค่า"}
_ORIGIN_KEYWORDS = {"origin", "country of origin", "ถิ่นกำเนิด"}


def _find_header_row(rows: list) -> tuple[int, dict]:
    """หา row index ของ header และ column mapping"""
    for ri, row in enumerate(rows):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        desc_col = next((ci for ci, v in enumerate(row_lower)
                         if any(kw in v for kw in _ITEM_HEADER_KEYWORDS)), None)
        if desc_col is not None:
            col_map = {"desc": desc_col}
            for ci, v in enumerate(row_lower):
                if any(kw in v for kw in _HS_KEYWORDS): col_map["hs"] = ci
                if any(kw in v for kw in _QTY_KEYWORDS) and "qty" not in col_map: col_map["qty"] = ci
                if any(kw in v for kw in _PRICE_KEYWORDS): col_map["price"] = ci
                if any(kw in v for kw in _AMOUNT_KEYWORDS): col_map["amount"] = ci
                if any(kw in v for kw in _ORIGIN_KEYWORDS): col_map["origin"] = ci
            return ri, col_map
    return -1, {}


def _safe_float(v):
    try:
        return float(str(v).replace(",", "").strip()) if v not in (None, "", "None") else None
    except Exception:
        return None


def _smart_parse_excel(content: bytes, filename: str) -> dict:
    """
    อ่าน Excel โดยตรงด้วย openpyxl
    - ค้นหา header row (description/goods/item)
    - extract items จาก rows ที่ตามมา
    - สกัด invoice meta จาก rows บนๆ (invoice_no, seller, buyer)
    """
    result = {"items": []}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]

        all_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        # สกัด meta จาก rows แรก (scan 30 rows แรก)
        meta_text = ""
        for row in all_rows[:30]:
            vals = [str(v).strip() for v in row if v not in (None, "")]
            if vals:
                meta_text += " | ".join(vals) + "\n"

        # ดึง invoice_no, seller, buyer จาก meta_text แบบ heuristic
        import re
        inv_match = re.search(r'INV[-][w\-]+', meta_text, re.IGNORECASE)
        result["invoice_no"] = inv_match.group(0) if inv_match else None

        date_match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2})', meta_text)
        result["invoice_date"] = date_match.group(1) if date_match else None

        # incoterms
        inco_match = re.search(r'\b(CIF|FOB|EXW|DDP|CFR|DAP|FCA)\b', meta_text, re.IGNORECASE)
        result["incoterms"] = inco_match.group(1).upper() if inco_match else None

        # currency
        cur_match = re.search(r'\b(USD|THB|CNY|EUR|GBP|JPY)\b', meta_text, re.IGNORECASE)
        result["currency"] = cur_match.group(1).upper() if cur_match else "USD"

        # seller/buyer country
        result["seller_country"] = "CN"
        result["buyer_country"] = "TH"
        if "china" in meta_text.lower(): result["seller_country"] = "CN"
        if "thailand" in meta_text.lower() or "bangkok" in meta_text.lower():
            result["buyer_country"] = "TH"

        # Seller/Buyer name (first non-empty text after keyword)
        for line in meta_text.split("\n"):
            if "seller" in line.lower() or "exporter" in line.lower() or "shipper" in line.lower():
                parts = [p.strip() for p in line.split("|") if p.strip()]
                if len(parts) > 1:
                    result["seller_name"] = parts[1]
                    break
        for line in meta_text.split("\n"):
            if "buyer" in line.lower() or "consignee" in line.lower() or "importer" in line.lower():
                parts = [p.strip() for p in line.split("|") if p.strip()]
                if len(parts) > 1:
                    result["buyer_name"] = parts[1]
                    break

        # หา header row
        header_ri, col_map = _find_header_row(all_rows)
        result["_debug"] = f"total_rows={len(all_rows)} header_ri={header_ri} col_map={col_map}"
        if header_ri >= 0:
            result["_debug"] += f" | header={all_rows[header_ri]}"
        if header_ri == -1 or "desc" not in col_map:
            result["_debug"] += " | rows_sample=" + str([str(r[:3]) for r in all_rows[:15] if any(v for v in r if v)])
            return result  # ไม่เจอ table → fallback Claude

        # Extract items จาก rows หลัง header
        items = []
        for row in all_rows[header_ri + 1:]:
            desc_val = row[col_map["desc"]] if col_map["desc"] < len(row) else None
            if not desc_val or str(desc_val).strip() in ("", "None"):
                continue
            desc_str = str(desc_val).strip()
            # ข้าม footer rows (TOTAL, SUBTOTAL, ฯลฯ)
            if any(kw in desc_str.upper() for kw in ("TOTAL", "SUBTOTAL", "REMARK", "SIGNATURE", "FREIGHT", "INSURANCE")):
                continue

            item = {
                "line_no": len(items) + 1,
                "description": desc_str,
                "hs_code_declared": str(row[col_map["hs"]]).strip() if "hs" in col_map and row[col_map["hs"]] else None,
                "qty": _safe_float(row[col_map["qty"]]) if "qty" in col_map else None,
                "unit": None,
                "unit_price": _safe_float(row[col_map["price"]]) if "price" in col_map else None,
                "line_value": _safe_float(row[col_map["amount"]]) if "amount" in col_map else None,
                "country_origin": str(row[col_map["origin"]]).strip() if "origin" in col_map and row[col_map["origin"]] else result.get("seller_country"),
                "marks_numbers": None,
            }
            # หา unit column (ถัดจาก qty)
            if "qty" in col_map:
                unit_ci = col_map["qty"] + 1
                if unit_ci < len(row) and unit_ci not in col_map.values():
                    item["unit"] = str(row[unit_ci]).strip() if row[unit_ci] else None

            items.append(item)

        result["items"] = items
    except Exception as e:
        result["_smart_parse_error"] = str(e)
    return result


# ─── Main parse function ──────────────────────────────────────────────────────

async def parse_invoice(filename: str, content: bytes) -> dict:
    """
    Entry point — ส่ง filename + bytes → คืน structured dict พร้อม raw_text
    {
      file_type, raw_text,
      invoice_no, invoice_date,
      seller_name, seller_country,
      buyer_name, buyer_country,
      incoterms, currency,
      items: [ {...} ]
    }
    """
    file_type = detect_file_type(filename, content)
    raw_text = ""
    extracted = {}

    if file_type == "PDF_TEXT":
        raw_text = extract_text_from_pdf(content)
        extracted = await extract_with_claude_text(raw_text)

    elif file_type in ("PDF_SCAN", "IMAGE"):
        extracted = await extract_with_claude_vision(content, file_type)
        raw_text = json.dumps(extracted, ensure_ascii=False)

    elif file_type in ("EXCEL", "CSV"):
        raw_text = ""
        # 1. Smart parse: อ่าน cells โดยตรง (ไม่พึ่ง text)
        if file_type == "EXCEL":
            extracted = _smart_parse_excel(content, filename)
            if extracted.get("items"):
                raw_text = _compile_excel_text(content, filename)
            else:
                extracted = {"items": []}

        # 2. Fallback: compile → text parse
        if not extracted.get("items"):
            raw_text = _compile_excel_text(content, filename)
            extracted = _parse_compiled_text(raw_text)

        # 3. Fallback: Claude (กรณี format แปลก)
        if not extracted.get("items"):
            if not raw_text:
                raw_text = _compile_excel_text(content, filename)
            extracted = await extract_with_claude_text(raw_text)

    else:
        return {"error": f"Unsupported file type: {filename}", "file_type": "UNKNOWN"}

    if "error" in extracted:
        return {"error": extracted["error"], "file_type": file_type, "raw_text": raw_text}

    extracted["file_type"] = file_type
    extracted["raw_text"] = raw_text[:5000]
    return extracted
